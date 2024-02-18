import logging
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from operators.dataframe_operator import DataFrameOperator

logger = logging.getLogger(__name__)


class ExcelOperator:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._load_workbook()

    def _load_workbook(self) -> None:
        """Ensure the Excel file exists; if not, create a new one."""
        if not os.path.exists(self.file_path):
            # Create a new Excel file with a default sheet
            workbook = Workbook()
            workbook.save(self.file_path)
        self._workbook = load_workbook(self.file_path)

    def _validate_worksheet(self, sheet_name: str, expected_columns: list[str]) -> None:
        if sheet_name not in self._workbook.sheetnames:
            # If the sheet does not exist, create it and add the expected columns as headers
            sheet = self._workbook.create_sheet(sheet_name)
            sheet.append(expected_columns)
        else:
            sheet = self._workbook[sheet_name]
            existing_columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            # Check if the existing columns in the sheet match the expected columns
            if set(existing_columns) != set(expected_columns):
                # Log a warning or raise an error as per your application's requirements
                logger.warning(f"Column mismatch in sheet '{sheet_name}'")
                logger.warning(f"[{sheet_name}] Existing columns: {existing_columns}")
                logger.warning(f"[{sheet_name}] Expected columns: {expected_columns}")

        self._workbook.save(self.file_path)

    def update_rows(self, sheet_name: str, rows_to_update: pd.DataFrame, model_class) -> None:
        sheet = self._workbook[sheet_name]

        # Read the headers from the Excel sheet
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        # Get the identifying fields from the model class
        identifying_fields = model_class.IDENTIFYING_FIELDS

        # Create a mapping of composite keys to Excel row numbers
        composite_key_to_row = {}
        for excel_row in range(2, sheet.max_row + 1):  # Start from 2 to skip the header row
            excel_row_key = tuple(sheet.cell(row=excel_row, column=headers.index(field) + 1).value for field in identifying_fields if field in headers)
            composite_key_to_row[excel_row_key] = excel_row

        # Iterate over the DataFrame rows to update the Excel sheet
        for _, row in rows_to_update.iterrows():
            # Construct the composite key for the current DataFrame row
            composite_key = tuple(row[field] for field in identifying_fields if field in row)

            # Use the mapping to find the Excel row that matches this composite key
            if composite_key in composite_key_to_row:
                excel_row = composite_key_to_row[composite_key]
                logger.debug(f"Matching row found in Excel at row {excel_row} for composite key {composite_key}")

                # Update the cells in this row with new values from the DataFrame
                for header in headers:
                    if header in row.index:  # Ensure the header is present in the DataFrame row
                        cell = sheet.cell(row=excel_row, column=headers.index(header) + 1)
                        cell.value = row[header]
            else:
                logger.warning(f"No matching row found in Excel for composite key {composite_key}")

    def append_rows(self, sheet_name: str, rows_to_append: pd.DataFrame) -> None:
        """Append new rows to the Excel sheet ensuring DataFrame columns are rearranged to match the sheet's columns."""
        sheet = self._workbook[sheet_name]

        existing_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        # Reorder DataFrame columns to match the sheet's headers, adding missing columns as empty if necessary
        for header in existing_headers:
            if header not in rows_to_append.columns:
                rows_to_append[header] = ""  # Add missing column as empty
        rows_to_append = rows_to_append[existing_headers]  # Reorder columns to match the sheet

        # Append rows from the DataFrame
        for _, row in rows_to_append.iterrows():
            sheet.append(row.tolist())

    def delete_rows(self, sheet_name: str, rows_to_delete: pd.DataFrame, model_class) -> None:
        """Cross out rows in the Excel sheet identified for deletion."""
        sheet = self._workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        identifying_fields = model_class.IDENTIFYING_FIELDS

        # Iterate over the DataFrame rows to find and cross out the corresponding rows in the Excel sheet
        for _, row in rows_to_delete.iterrows():
            composite_key = tuple(row[field] for field in identifying_fields if field in row)

            for excel_row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header row
                excel_row_key = tuple(sheet.cell(row=excel_row, column=headers.index(field) + 1).value for field in identifying_fields if field in headers)

                if excel_row_key == composite_key:
                    # Apply strikethrough style to each cell in this row
                    for col_index in range(1, len(headers) + 1):
                        cell = sheet.cell(row=excel_row, column=col_index)
                        cell.font = Font(strike=True)

                    break  # Break the loop once the matching row is found and processed

    def push_updates(self, sheet_name: str, new_data: pd.DataFrame, model_class) -> None:
        # Validate the worksheet structure
        self._validate_worksheet(sheet_name, model_class.objects.get_field_names())
        new_data = DataFrameOperator.remove_timezone(new_data)
        old_data = pd.read_excel(self.file_path, sheet_name=sheet_name)

        # Define unique fields for merging; these should be the identifying fields of your model
        unique_fields = model_class.IDENTIFYING_FIELDS

        # If the sheet exists and has data, categorize updates
        if not old_data.empty:
            # Rename columns in new_data to avoid suffix conflicts during merge
            new_data_renamed = new_data.rename(columns={col: f"{col}_new" for col in new_data.columns if col not in unique_fields})
            old_data_renamed = old_data.rename(columns={col: f"{col}_old" for col in old_data.columns if col not in unique_fields})

            # Merge data with an indicator to determine the source of each row
            combined_data = pd.merge(old_data_renamed, new_data_renamed, on=unique_fields, how="outer", indicator=True)

            # Categorize rows based on the _merge indicator
            rows_to_update = combined_data[combined_data["_merge"] == "both"]
            rows_to_append = combined_data[combined_data["_merge"] == "right_only"]
            rows_to_delete = combined_data[combined_data["_merge"] == "left_only"]

            # Clean up rows_to_update: remove '_old' suffix and '_merge' column
            rows_to_update = rows_to_update.rename(columns=lambda col: col.replace("_new", ""))
            rows_to_update.drop(columns=["_merge"] + [col for col in rows_to_update if col.endswith("_old")], inplace=True)

            # Clean up rows_to_append: remove '_new' suffix and '_merge' column
            rows_to_append = rows_to_append.rename(columns=lambda col: col.replace("_new", ""))
            rows_to_append.drop(columns=["_merge"] + [col for col in rows_to_append if col.endswith("_old")], inplace=True)

            # Clean up rows_to_delete: drop '_merge' column
            rows_to_delete = rows_to_delete.drop(columns=["_merge"])
        else:
            # If there's no existing data, all rows from the new data are to be appended
            rows_to_append = new_data
            rows_to_update = pd.DataFrame()
            rows_to_delete = pd.DataFrame()

        # Now, call the new methods with the categorized rows
        if not rows_to_update.empty:
            rows_to_update.reset_index(drop=True, inplace=True)
            self.update_rows(sheet_name, rows_to_update, model_class)

        if not rows_to_append.empty:
            rows_to_append.reset_index(drop=True, inplace=True)
            self.append_rows(sheet_name, rows_to_append)

        if not rows_to_delete.empty:
            self.delete_rows(sheet_name, rows_to_delete, model_class)

        self._workbook.save(self.file_path)

    def pull_updates(self, sheet_name: str) -> pd.DataFrame:
        """Read data from an Excel sheet and return it as a DataFrame."""
        sheet = self._workbook[sheet_name]
        data = pd.DataFrame(sheet.values)
        data.columns = data.iloc[0]
        data = data.iloc[1:]
        return data
